import json
import math
import openpyxl
from pathlib import Path

# ClientProperties
directoryId = ""

# Tags
user_key = ""
user_value = ""

# Process config
names = []
users = {}
ready_list = []
workspacesId = []
data = []

# Pre data processing
# Load user database
xlsx_users = Path('', 'aws_workspaces_database.xlsx')
client_users = openpyxl.load_workbook(xlsx_users)
user_sheet = client_users.active

# Append users in database into users{}:
for row in user_sheet.iter_rows(2):
    cell_values = []
    for cell in row:
        cell_value = cell.value

        cell_values.append(cell.value)
    users[cell_values[1]] = cell_values[0]
    workspacesId.append(cell_values[0])
    cell_values.clear()

# Load users to remove
xlsx_to_delete = Path('', 'users_to_delete.xlsx')
users_to_delete = openpyxl.load_workbook(xlsx_to_delete)
delete_sheet = users_to_delete.active
n_jsons = math.ceil(delete_sheet.max_row / 25)

# Append users to remove into names[]:
for row in delete_sheet.iter_rows(delete_sheet.max_column):
    for cell in row:
        cell_value = cell.value
        names.append(cell_value)
user_count = len(names)
print(names)


# Create JSON files function
def process():
    aws_max_workload = 25
    for g in range(n_jsons):
        if aws_max_workload > len(names):
            aws_max_workload = len(names)
        for h in range(aws_max_workload):
            data.append({
                'WorkspaceId': users[names[h]],
            })

        with open('data' + str(g) + '.json', 'w') as outfile:
            json.dump(data, outfile)

        for index in range(aws_max_workload):
            names.pop(0)

        data.clear()

    return True


# UI Start
print()

print("***aws_wsAutoTerminateV1.0***")
print(".-._.-._._.W311C0M3._._.-._.-.")
print("***aws_wsAutoTerminateV1.0***")

print()
print()

print("You are about to terminate " + str(user_count) + " workspaces.")
print()


if input("Apply changes?(y/n) ") == "y":
    print("Creating .json files...")
    process()
    print("*********************************************")
    print("************ OPERATION SUCCESS **************")
    print("*********************************************")
    print("************* OUTPUT " + str(n_jsons) + " .JSONs ***************")
    print("*********************************************")
    print("************ OPERACION EXITOSA **************")
    print("*********************************************")
else:
    print("Operation cancelled. Rerun script to continue...")
    exit()





